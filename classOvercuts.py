import pyautogui as pt
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import regex
import shutil
from time import sleep


class Overcuts:
    def __int__(self):
        # Locations of the Required Files
        self.LOCATION = os.path.abspath(os.path.dirname(__file__))  # Script Folder
        self.excel_file_original = self.LOCATION + r"\data\overcuts-v2.xlsx"  # Original Excel File
        self.excel_file = self.LOCATION + r"\data\overcuts-temp.xlsx"  # temp excel file
        self.img_distance_field = self.LOCATION + r"\data\distance_field_laptop.png"  # Distance Field image to location
        self.pytesseract = pytesseract
        self.pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\Tesseract.exe"  # Tesseract OCR
        self.img1 = "data/dd1.png"  # Template image - 1 to store new created image from screenshot
        self.img2 = "data/dd2.png"  # Template image - 2
        self.img3 = "data/dd3.png"  # Template image - 3
        self.img4 = "data/dd4.png"  # Template image - 4
        self.row = 0  # Starting Row Number for Excel
        self.col = 0  # Starting Column Number for Excel
        self.required_files = [self.excel_file_original, self.pytesseract.pytesseract.tesseract_cmd, self.img_distance_field]

        # Checks the Required files for Script
    def scan_required_data(self):
        for f in self.required_files:
            try:
                with open(f, "r") as file:
                    pass    # if file found
            except FileNotFoundError as e:
                print(f"Error: File not Found\n"    # file not found
                      f"{f}")

    # Function to start an Excel file which will save all the required overcuts data from FarmWorks
    def start_excel(self):
        try:
            os.remove(self.excel_file)  # Remove Existing Temp File
            shutil.copy(self.excel_file_original, self.excel_file)  # Create New Temp File
        except PermissionError:
            print("Error: The Excel template file is already open or in use by another program."
                  "Please Close the Excel File and Start program again.")
            exit()

        self.oxl = load_workbook(self.excel_file)
        self.xl_sheet = self.oxl['M1']  # opening M1 Sheet
        self.row = 6    # starting row

    # Adds values to Excel files at provided Row and Column Location
    def add_to_excel(self, input_cell, input_value):
        try:
            self.xl_sheet[input_cell] = input_value
        except Exception as e:
            print(f"""Error : {type(e).__name__}
                        Location: Entering Data into Excel File""")
            exit()

    # Close Excel File
    def close_excel(self):
        # Formula to Remove Empty Rows. Did not work as expected
        '''self.xl_sheet.delete_rows(self.row, 41 - self.row)
        for row in self.xl_sheet.iter_rows(min_row=self.row, max_row=40):
            for cell in row:
                cell.value = None'''
        # Function to resize table - NOT WORKING
        '''table = self.xl_sheet.tables['Table1']
        end_row = 'L' + str(self.row)
        table.ref = 'A1:' + end_row'''
        try:
            self.oxl.save(self.excel_file)
        except FileNotFoundError as e:
            print(f"Error: File not Found\n"  # file not found
                  f"{self.excel_file}")

    # Open Temp Excel file
    def show_result(self):
        try:
            os.startfile(self.excel_file)
        except FileNotFoundError as e:
            print(f"Error: File not Found\n"  # file not found
                  f"{self.excel_file}")

    # This function is used to find the location of an image on screen
    @staticmethod
    def locate_image(image_location, probability=0.7):
        # Arguments: image_location=Location of image that to be located,
        # probability=Similarity of image to be located to the image found on desktop, Value=0 to 1
        try:
            position = pt.locateOnScreen(image_location, confidence=probability)
            x = position[0]
            y = position[1]
            return x, y
        except Exception as e:  # if image not found
            print(f"""Error : {type(e).__name__}
            Location: Locating Image \n{image_location}""")
            exit()

    # This function locates the image of distance field using locate_image function
    def locate_distance_field(self):
        coordinates = self.locate_image(self.img_distance_field, 0.7)
        self.x_coordinate = coordinates[0]
        self.y_coordinate = coordinates[1]

        # Try and except case in case mouse cursor is unable to move to the coordinates location
        try:
            pt.moveTo(self.x_coordinate, self.y_coordinate, duration=.05)
        except Exception as e:
            print(f"""Error : {type(e).__name__}
            Location: Moving Mouse Cursor to Distance Field""")
            exit()

    # This function moves the cursor to Entered distance value
    def move_to_distance(self, input_distance):
        pt.doubleClick(self.x_coordinate + 110, self.y_coordinate + 12, interval=0)
        pt.press('backspace')

        # Converting input_distance integer to string to enter each character one by one into the input field
        text_to_str = str(input_distance)
        k = 0
        while k < len(str(text_to_str)):
            pt.press(text_to_str[k])
            k += 1
        sleep(0.3)

    # Take a snip out of screen and converts that snip of image into text using Tesseract OCR
    def snip_to_text(self, configuration, add_to_x=0, add_to_y=0, image_size_x=67, image_size_y=17, output_type="text",
                     print_text=""):
        try:
            im = pt.screenshot(region=(self.x_coordinate + add_to_x, self.y_coordinate - add_to_y, image_size_x, image_size_y))
            im.save(self.img1)
            text = pytesseract.image_to_string(Image.open(self.img1), lang="eng", config=configuration)
            # Test - print("OCR: " + text)
        except Exception as e:
            print(f"""Error: {type(e).__name__}
            Location: OCR""")
            exit()

        if not text:
            print(f"Error: No Text Found During OCR\nWe might have reached"
                  f" at the end of Tile.")
            self.close_excel()
            self.show_result()
            exit()

        if output_type == "number":
            # Cleaning Text - Removing White Spaces and Non-Alphanumeric Symbols
            # print("Cleaning Number: " + text + "\n")
            text = text.replace('O', '0').replace('o', '0').replace('\n', '')
            try:
                text = float(regex.sub(r"([^\d.]|(?<=\..*)\.)", "", text))
            except Exception as e:
                print(f"""Error: {type(e).__name__}
                    Location: Cleaning Number - {text}""")
                self.close_excel()
                self.show_result()
                exit()

            if print_text:
                print(f"{print_text}: " + str(text))
        #sleep(0.5)
        return text


# User Inputs (Start and End) to start Overcut Script
starting_distance = int(input('Enter Starting Point for Overcut: '))
stop_distance = int(input('Enter End Point for Overcut: '))
error_trials = 0

# Starting Overcut
print("------------Program Starts------------\n")
overcut_start = Overcuts()
overcut_start.__int__()
overcut_start.start_excel()
overcut_start.locate_distance_field()
number_format = {'num_format': '#,##0.00'}
configuration = ' --psm 7 -c tessedit_char_whitelist=0123456789.'
while starting_distance <= stop_distance:
    print(f"Dist: {starting_distance}")
    overcut_start.move_to_distance(starting_distance)
    overcut_start.add_to_excel("A" + str(overcut_start.row), starting_distance)  # Starting Distance to Excel
    surface_val = overcut_start.snip_to_text(configuration, 66, 95, 67, 17, "number", print_text="Surface")
    overcut_start.add_to_excel("F" + str(overcut_start.row), surface_val)  # Surface Value to Excel
    elevation_val = overcut_start.snip_to_text(configuration, 66, 71, 67, 17, "number", print_text="Elevation")
    overcut_start.add_to_excel("G" + str(overcut_start.row), elevation_val)  # Elevation Value to Excel
    depth_val = overcut_start.snip_to_text(configuration, 66, 49, 53, 19, "number", print_text="Captured Depth")
    overcut_start.add_to_excel("N" + str(overcut_start.row), depth_val)  # Depth Value to Excel
    tile_size_val = overcut_start.snip_to_text(configuration, 66, 119, 56, 17, "number", print_text="Tile Size")
    overcut_start.add_to_excel("C" + str(overcut_start.row), tile_size_val)  # Tile Size Value to Excel

    # Depth Calculation to verify it with captured depth
    calculated_depth = round(float(surface_val - elevation_val), 2)
    
    if depth_val - calculated_depth > 1:
        if error_trials < 2:
            print(f"****Calculation Error****\n"
                  f"Captured Depth({depth_val}) is not equal to Calculated Depth({calculated_depth}).\n"
                  f"Retrying in 2 seconds...")
            if error_trials == 0:
                print("Max Tries = 3")
            print("")
            overcut_start.row -= 1
            starting_distance = starting_distance - 100
            error_trials += 1
            sleep(2)
        else:
            print(f"Failed to Capture Correct value for Depth at Distance: {starting_distance}"
                  f"\n-----------------------\n")
            overcut_start.add_to_excel("O" + str(overcut_start.row), calculated_depth)
            red_font = Font(color='FF0000', bold=True)
            overcut_start.xl_sheet["O" + str(overcut_start.row)].font = red_font
    else:
        overcut_start.add_to_excel("O" + str(overcut_start.row), calculated_depth)
        error_trials = 0
        print("\n-----------------------\n")

    overcut_start.row += 1
    starting_distance = starting_distance + 100

    # End of Loop when Start Distance is > End Distance

overcut_start.close_excel()
overcut_start.show_result()
